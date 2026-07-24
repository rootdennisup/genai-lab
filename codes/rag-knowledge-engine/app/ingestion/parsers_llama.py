"""基于 LlamaIndex Reader 的文档解析器（学习参考实现）。

这个模块与 :mod:`app.ingestion.parsers` 提供相同的公开入口和输出类型：

``文件路径 -> 格式 Reader -> LlamaIndex Document -> ParsedBlock``

设计上只让 LlamaIndex 负责“读取文件”。Chunk 切分仍由项目自己的 ``chunker.py``
完成，避免 Reader、NodeParser 和业务 Chunk 策略混在一起。这样既能学习 LlamaIndex
的 Data Connector 抽象，又不会让领域层依赖 LlamaIndex 的 ``Document`` 类型。

安装可选依赖：``pip install -e ".[llama]"``。
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any, Protocol

from app.domain.models import ParsedBlock


class ParseError(ValueError):
    """文件无法被 Reader 可靠解析时抛出的领域错误。"""


class LlamaDocument(Protocol):
    """本模块真正使用到的 LlamaIndex Document 最小接口。

    使用 Protocol 而不是把 LlamaIndex 类型传入领域层，有两个好处：
    1. 领域模型继续与第三方框架解耦；
    2. 映射逻辑可以用很小的假对象做单元测试。
    """

    metadata: dict[str, Any]

    def get_content(self) -> str: ...


def parse_document(path: Path, file_type: str) -> list[ParsedBlock]:
    """使用指定 Reader 解析文件，并统一转换成 ``ParsedBlock``。

    这个函数可以作为原 ``parsers.parse_document`` 的替代品调用。这里显式接收
    ``file_type``，而不是仅相信扩展名，保持与现有摄取服务的接口完全一致。
    """

    parsers = {
        "md": _parse_markdown,
        "pdf": _parse_pdf,
        "docx": _parse_docx,
        "xlsx": _parse_xlsx,
    }
    try:
        parser = parsers[file_type]
    except KeyError as exc:
        raise ParseError(f"不支持的文件类型：{file_type}") from exc

    if not path.is_file():
        raise ParseError(f"待解析文件不存在：{path.name}")

    try:
        blocks = parser(path)
    except ParseError:
        raise
    except Exception as exc:
        # API 层不应收到第三方包的完整堆栈、临时路径或其他内部信息。
        raise ParseError(f"{file_type.upper()} 文档解析失败：{exc}") from exc

    if not blocks:
        raise ParseError("文档未提取到可索引文本")
    return blocks


def _reader_types() -> tuple[type, type, type, type]:
    """延迟导入可选依赖，使主项目不安装 LlamaIndex 时仍可正常启动。

    LlamaIndex 采用拆包设计：核心 ``Document`` 在 ``llama-index-core``，文件 Reader
    在 ``llama-index-readers-file``。Reader 应从 ``llama_index.readers.file`` 导入，
    而不是旧版本曾使用的 ``llama_index.core.readers.file``。
    """

    try:
        from llama_index.readers.file import (
            DocxReader,
            MarkdownReader,
            PandasExcelReader,
            PDFReader,
        )
    except ImportError as exc:
        raise ParseError(
            '使用 LlamaIndex 解析器需要安装可选依赖：pip install -e ".[llama]"'
        ) from exc
    return PDFReader, DocxReader, MarkdownReader, PandasExcelReader


def _clean(text: str) -> str:
    """Reader 之后仍做一次保守清洗，但不删除标点、编号和专有名词。"""

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _content(document: LlamaDocument) -> str:
    """兼容 LlamaIndex 新旧 Document 的正文访问方式。"""

    get_content = getattr(document, "get_content", None)
    if callable(get_content):
        return _clean(str(get_content()))
    # ``text`` 是部分旧版本 Document 的公开属性，仅作为学习示例的兼容降级。
    return _clean(str(getattr(document, "text", "")))


def _metadata(document: LlamaDocument) -> dict[str, Any]:
    metadata = getattr(document, "metadata", None)
    return dict(metadata) if isinstance(metadata, dict) else {}


def _first(metadata: dict[str, Any], *keys: str) -> Any | None:
    """Reader 版本间 metadata 键名偶有变化，按候选名称读取第一个有效值。"""

    for key in keys:
        value = metadata.get(key)
        if value not in (None, ""):
            return value
    return None


def _as_positive_int(value: Any, fallback: int) -> int:
    try:
        parsed = int(value)
        return parsed if parsed > 0 else fallback
    except (TypeError, ValueError):
        return fallback


def _documents_to_blocks(
    documents: Iterable[LlamaDocument],
    path: Path,
    *,
    locator_factory: Any | None = None,
) -> list[ParsedBlock]:
    """执行框架对象到领域对象的防腐层映射。

    ``locator_factory`` 用于不同格式补充可靠定位信息。不能直接把 Reader 的全部
    metadata 原样保存，因为其中可能包含绝对路径，也可能随版本变化。
    """

    blocks: list[ParsedBlock] = []
    for index, document in enumerate(documents, 1):
        text = _content(document)
        if not text:
            continue
        metadata = _metadata(document)
        title = str(_first(metadata, "title", "section", "file_name") or path.stem)
        locator = locator_factory(metadata, index) if locator_factory else {"document_part": index}
        blocks.append(
            ParsedBlock(
                text=text,
                title=title,
                section_path=[title] if title != path.stem else [],
                locator=locator,
            )
        )
    return blocks


def _parse_pdf(path: Path) -> list[ParsedBlock]:
    """PDFReader 默认按页返回 Document，因此每页可形成一个可引用结构块。"""

    PDFReader, _, _, _ = _reader_types()
    documents = PDFReader(return_full_document=False).load_data(file=path)

    def page_locator(metadata: dict[str, Any], index: int) -> dict[str, int]:
        # 常见键为 page_label；index 是 Reader 未提供页码时的安全后备值。
        page = _as_positive_int(_first(metadata, "page_label", "page_number", "page"), index)
        return {"page_start": page, "page_end": page}

    blocks = _documents_to_blocks(documents, path, locator_factory=page_locator)
    if not blocks:
        raise ParseError("PDF 未提取到文本，可能是扫描件；此参考实现不包含 OCR")
    return blocks


def _parse_docx(path: Path) -> list[ParsedBlock]:
    """DocxReader 读取 Word 正文；通常返回整篇或较粗粒度的 Document。

    与自研解析器相比，这个 Reader 抽象更简单，但默认不保证保留标题样式、段落序号
    和表格行坐标。这里如实保存 ``document_part``，不伪造页码或章节位置。
    """

    _, DocxReader, _, _ = _reader_types()
    documents = DocxReader().load_data(file=path)
    return _documents_to_blocks(documents, path)


def _parse_markdown(path: Path) -> list[ParsedBlock]:
    """MarkdownReader 负责读取 Markdown，不在 Reader 阶段提前执行 Chunk。"""

    _, _, MarkdownReader, _ = _reader_types()
    documents = MarkdownReader().load_data(file=path)
    blocks = _documents_to_blocks(documents, path)

    # Reader 通常不会提供源文件行号。为了引用可追溯，只用原文精确匹配补回范围；
    # 若 Reader 改写了文本导致无法精确匹配，则明确退化为 document_part。
    source_lines = path.read_text(encoding="utf-8-sig").splitlines()
    source_text = "\n".join(source_lines)
    for block in blocks:
        start = source_text.find(block.text)
        if start >= 0:
            line_start = source_text.count("\n", 0, start) + 1
            block.locator = {
                "line_start": line_start,
                "line_end": line_start + block.text.count("\n"),
            }
    return blocks


def _parse_xlsx(path: Path) -> list[ParsedBlock]:
    """按工作表调用 PandasExcelReader，并把工作表名写入来源定位。

    ``concat_rows=True`` 会把一个工作表聚合成一个 Document，适合演示 Reader；它不
    能像原解析器一样可靠保留每条数据的行号，因此 locator 只声明工作表，不伪造行号。
    如果问答要求精确到 Excel 行，原 ``parsers.py`` 的逐行实现更合适。
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParseError("读取 Excel 工作表名称需要安装 openpyxl") from exc

    _, _, _, PandasExcelReader = _reader_types()
    # 这里只用 openpyxl 获取 sheet 名；正文解析完全交给 LlamaIndex Reader。
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    blocks: list[ParsedBlock] = []
    for sheet_name in sheet_names:
        reader = PandasExcelReader(concat_rows=True, sheet_name=sheet_name)
        documents = reader.load_data(file=path, extra_info={"sheet_name": sheet_name})
        for block in _documents_to_blocks(
            documents,
            path,
            locator_factory=lambda _metadata, _index, sheet=sheet_name: {"sheet_name": sheet},
        ):
            block.title = sheet_name
            block.section_path = [sheet_name]
            blocks.append(block)
    return blocks

